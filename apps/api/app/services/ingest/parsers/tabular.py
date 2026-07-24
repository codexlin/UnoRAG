"""Native CSV/XLSX parsers producing one canonical TableIR per table/sheet."""

from __future__ import annotations

import csv
import io
from datetime import date, datetime
from uuid import uuid4

from app.services.ingest.ir import (
	DocumentIR,
	Node,
	NodeType,
	ParserReport,
	content_hash_bytes,
)
from app.services.ingest.parsers.txt import decode_text_bytes
from app.services.ingest.table_ir import TableIR, normalize_table, table_ir_to_legacy


def parse_csv(
	*,
	content: bytes,
	filename: str,
	title: str,
	doc_id: str | None = None,
	library_id: str = "",
) -> DocumentIR:
	text = decode_text_bytes(content)
	try:
		dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
	except csv.Error:
		dialect = csv.excel
	rows = [
		[_cell_text(value) for value in row]
		for row in csv.reader(io.StringIO(text), dialect)
		if any(str(value).strip() for value in row)
	]
	if not rows:
		raise ValueError("CSV produced no rows")
	table_id = "csv-t1"
	table_ir = normalize_table(
		table_id=table_id,
		headers=rows[0],
		rows=rows[1:],
		caption=title,
		confidence=0.98,
	)
	node = _table_node(table_ir, path=title, confidence=0.98)
	return DocumentIR(
		id=doc_id or str(uuid4()),
		library_id=library_id,
		source=filename,
		source_format="csv",
		title=title,
		filename=filename,
		content_hash=content_hash_bytes(content),
		nodes=[node],
		parser_report=ParserReport(
			source_format="csv",
			parser="python_csv",
			backend="stdlib",
			mode="structured",
			metrics={"node_count": 1, "table_count": 1},
			notes="native CSV rows preserved",
		),
	)


def parse_xlsx(
	*,
	content: bytes,
	filename: str,
	title: str,
	doc_id: str | None = None,
	library_id: str = "",
) -> DocumentIR:
	try:
		from openpyxl import load_workbook
	except ImportError as exc:
		raise ValueError("XLSX support requires openpyxl") from exc

	workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
	nodes: list[Node] = []
	table_count = 0
	try:
		for sheet in workbook.worksheets:
			rows = [
				[_cell_text(value) for value in row]
				for row in sheet.iter_rows(values_only=True)
			]
			rows = [row for row in rows if any(value for value in row)]
			if not rows:
				continue
			width = max(len(row) for row in rows)
			rows = [[*row, *([""] * (width - len(row)))] for row in rows]
			table_count += 1
			table_id = f"xlsx-t{table_count}"
			table_ir = normalize_table(
				table_id=table_id,
				headers=rows[0],
				rows=rows[1:],
				caption=sheet.title,
				confidence=0.98,
			)
			nodes.append(_table_node(table_ir, path=sheet.title, confidence=0.98))
	finally:
		workbook.close()

	if not nodes:
		raise ValueError("XLSX produced no non-empty worksheets")
	return DocumentIR(
		id=doc_id or str(uuid4()),
		library_id=library_id,
		source=filename,
		source_format="xlsx",
		title=title,
		filename=filename,
		content_hash=content_hash_bytes(content),
		nodes=nodes,
		parser_report=ParserReport(
			source_format="xlsx",
			parser="openpyxl",
			backend="openpyxl",
			mode="structured",
			metrics={"node_count": len(nodes), "table_count": len(nodes)},
			notes="one logical TableIR per non-empty worksheet",
		),
	)


def _table_node(table_ir: TableIR, *, path: str, confidence: float) -> Node:
	table = TableIR.model_validate(table_ir)
	lines = [" | ".join(table.headers())] if table.headers() else []
	lines.extend(" | ".join(row) for row in table.legacy_rows())
	return Node(
		id=str(uuid4()),
		type=NodeType.TABLE,
		path=path,
		text="\n".join(lines),
		table_json=table_ir_to_legacy(table),
		table_ir=table,
		table_id=table.table_id,
		confidence=confidence,
		meta={"caption": table.caption},
	)


def _cell_text(value: object) -> str:
	if value is None:
		return ""
	if isinstance(value, datetime):
		return value.isoformat(sep=" ", timespec="seconds")
	if isinstance(value, date):
		return value.isoformat()
	return str(value).strip()
